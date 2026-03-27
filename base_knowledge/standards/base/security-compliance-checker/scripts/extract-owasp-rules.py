#!/usr/bin/env python3
"""
OWASP Security Rule Extractor (Dynamic Format)

Đọc file Excel chứa OWASP rules, tự động phát hiện cấu trúc cột,
filter theo PIC, xuất owasp_checklist.md.

Script KHÔNG giả định tên file, tên sheet, hoặc tên cột cố định.
Mọi thứ được auto-detect tại runtime.

Usage:
    python extract-owasp-rules.py <path_to_xlsx> [options]

Options:
    --output, -o    Output file path (default: owasp_checklist.md cùng thư mục xlsx)
    --sheet, -s     Sheet name to read (default: first sheet)
    --pic           PIC values to filter (default: All BO)
    --list-sheets   List all sheet names and exit
    --preview       Preview headers and first 5 rows, then exit

Dependencies:
    pip install openpyxl
"""

import argparse
import sys
import os
import io
from datetime import datetime

# Fix Windows console encoding (cp1252 cannot handle Vietnamese Unicode)
if sys.stdout.encoding != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
if sys.stderr.encoding != 'utf-8':
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# Column Detection — Auto-detect column roles by fuzzy matching
# ============================================================

# Known patterns for each logical column role.
# Script will try to match column headers against these patterns.
COLUMN_PATTERNS = {
    'pic': ['pic'],
    'chapter_id': ['chapter_id', 'chapter id', 'chương', 'chapter'],
    'chapter_name': ['chapter_name', 'chapter name', 'tên chương'],
    'section_id': ['section_id', 'section id', 'mục'],
    'section_name': ['section_name', 'section name', 'tên mục'],
    'req_id': ['req_id', 'requirement_id', 'requirement id', 'id', 'mã yêu cầu', 'req id'],
    'req_description': ['req_description', 'requirement_description', 'description', 'mô tả', 'yêu cầu', 'requirement'],
    'l1': ['l1', 'level 1', 'level1', 'cấp 1'],
    'l2': ['l2', 'level 2', 'level2', 'cấp 2'],
    'l3': ['l3', 'level 3', 'level3', 'cấp 3'],
    'verify': ['verify', 'verification', 'kiểm tra', 'đáp ứng', 'xác minh'],
    'notes': ['notes', 'note', 'ghi chú', 'nhận xét', 'comment'],
}

# Minimum required columns to proceed
REQUIRED_ROLES = ['req_id', 'req_description']


def normalize_header(header_text):
    """Normalize a header string for matching."""
    if not header_text:
        return ''
    return str(header_text).strip().lower().replace('\n', ' ').replace('\r', '')


def detect_columns(headers_map):
    """
    Auto-detect which column corresponds to which logical role.

    Args:
        headers_map: dict {column_number: header_text}

    Returns:
        dict {role_name: column_number}
    """
    role_to_col = {}

    for role, patterns in COLUMN_PATTERNS.items():
        for col_num, header_text in headers_map.items():
            normalized = normalize_header(header_text)
            for pattern in patterns:
                # Short patterns (≤5 chars like 'pic', 'l1') require exact match
                # Longer patterns can use substring match
                if len(pattern) <= 5:
                    matched = (pattern == normalized)
                else:
                    matched = (pattern == normalized or pattern in normalized)

                if matched:
                    if role not in role_to_col:  # First match wins
                        role_to_col[role] = col_num
                    break

    return role_to_col


def find_header_row(ws, max_search=15):
    """
    Find the header row by looking for rows that contain
    column names matching known patterns.
    Requires at least 3 matched roles to avoid false positives on title rows.
    """
    best_row = None
    best_score = 0
    best_headers = None
    best_role_map = None

    for row_num in range(1, max_search + 1):
        headers = {}
        for cell in ws[row_num]:
            if not hasattr(cell, 'column') or cell.column is None:
                continue
            if cell.value:
                headers[cell.column] = str(cell.value).strip()

        if not headers:
            continue

        # Try to detect columns
        role_map = detect_columns(headers)
        score = len(role_map)

        # Track the row with the most matched roles
        if score > best_score:
            best_score = score
            best_row = row_num
            best_headers = headers
            best_role_map = role_map

    # Require at least 3 matched roles to avoid false positives on title rows
    if best_score >= 3:
        return best_row, best_headers, best_role_map

    return None, None, None


def extract_rules(xlsx_path, sheet_name=None, pic_filter=None):
    """
    Extract security rules from Excel file.
    Auto-detects column structure.

    Args:
        xlsx_path: Path to the Excel file
        sheet_name: Specific sheet to read (default: first sheet)
        pic_filter: List of PIC values to include (default: ['All', 'BO'])

    Returns:
        (rules_list, sheet_name, role_map, total_rows)
    """
    if pic_filter is None:
        pic_filter = ['ALL', 'BO']
    else:
        pic_filter = [p.upper() for p in pic_filter]

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            available = ', '.join(wb.sheetnames)
            print(f"ERROR: Sheet '{sheet_name}' not found.\nAvailable sheets: {available}", file=sys.stderr)
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]

    # Find header row and detect columns
    header_row, raw_headers, role_map = find_header_row(ws)
    if not header_row:
        print(f"ERROR: Could not auto-detect header row in sheet '{sheet_name}'.", file=sys.stderr)
        print(f"TIP: Use --preview to inspect the sheet structure.", file=sys.stderr)
        sys.exit(1)

    print(f"  Header row: {header_row}")
    print(f"  Detected columns:")
    for role, col_num in sorted(role_map.items(), key=lambda x: x[1]):
        print(f"    {role:20s} -> Col {col_num} ({raw_headers.get(col_num, '?')})")

    # Check required columns
    missing = [r for r in REQUIRED_ROLES if r not in role_map]
    if missing:
        print(f"\nERROR: Missing required columns: {missing}", file=sys.stderr)
        print(f"Detected: {list(role_map.keys())}", file=sys.stderr)
        sys.exit(1)

    # If no PIC column detected, include all rows
    has_pic = 'pic' in role_map

    # Build reverse map: column_number -> role_name
    col_to_role = {v: k for k, v in role_map.items()}

    # Extract data
    rules = []
    total_rows = 0
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        row_data = {}
        for cell in row:
            if not hasattr(cell, 'column') or cell.column is None:
                continue
            if cell.column in col_to_role:
                role = col_to_role[cell.column]
                row_data[role] = str(cell.value).strip() if cell.value is not None else ''

        req_id = row_data.get('req_id', '')
        if not req_id:
            continue

        total_rows += 1

        # Filter by PIC if column exists
        if has_pic:
            pic = row_data.get('pic', '')
            if not pic or pic.upper() not in pic_filter:
                continue
        else:
            row_data['pic'] = 'All'  # Default if no PIC column

        rules.append(row_data)

    wb.close()
    return rules, sheet_name, role_map, total_rows


def generate_markdown(rules, sheet_name, xlsx_path, role_map):
    """Generate structured markdown checklist from extracted rules."""
    lines = []
    lines.append("# OWASP ASVS Security Checklist")
    lines.append("")
    lines.append(f"> **Auto-generated** from `{os.path.basename(xlsx_path)}` (sheet: `{sheet_name}`)")
    lines.append(f"> Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"> Total rules: **{len(rules)}**")
    lines.append("")
    lines.append("---")
    lines.append("")

    has_chapter = 'chapter_id' in role_map
    has_section = 'section_id' in role_map
    has_levels = 'l1' in role_map

    current_chapter = ""
    current_section = ""

    for rule in rules:
        # Chapter grouping
        if has_chapter:
            chapter = rule.get('chapter_name', rule.get('chapter_id', ''))
            if chapter and chapter != current_chapter:
                current_chapter = chapter
                lines.append(f"## {rule.get('chapter_id', '')} — {chapter}")
                lines.append("")

        # Section grouping
        if has_section:
            section = rule.get('section_name', rule.get('section_id', ''))
            if section and section != current_section:
                current_section = section
                lines.append(f"### {rule.get('section_id', '')} — {section}")
                lines.append("")

                # Build table header dynamically
                header_cols = ["ID", "PIC"]
                if has_levels:
                    header_cols.extend(["L1", "L2", "L3"])
                header_cols.append("Description")
                if 'verify' in role_map:
                    header_cols.append("Verify")
                if 'notes' in role_map:
                    header_cols.append("Notes")
                lines.append("| " + " | ".join(header_cols) + " |")
                lines.append("|" + "|".join(["---"] * len(header_cols)) + "|")

        # Build row
        def escape(txt, max_len=200):
            return txt[:max_len].replace('|', '\\|').replace('\n', ' ') if txt else ''

        row_cols = [
            rule.get('req_id', ''),
            rule.get('pic', ''),
        ]

        if has_levels:
            for lvl in ['l1', 'l2', 'l3']:
                val = rule.get(lvl, '')
                row_cols.append('✅' if val and val.lower() == 'x' else '')

        row_cols.append(escape(rule.get('req_description', '')))

        if 'verify' in role_map:
            row_cols.append(escape(rule.get('verify', ''), 150))
        if 'notes' in role_map:
            row_cols.append(escape(rule.get('notes', ''), 150))

        lines.append("| " + " | ".join(row_cols) + " |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append(f"*End of checklist. Total: {len(rules)} rules.*")

    return '\n'.join(lines)


def list_sheets(xlsx_path):
    """List all sheet names in the Excel file."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    print(f"Sheets in '{os.path.basename(xlsx_path)}':")
    for i, name in enumerate(wb.sheetnames):
        print(f"  {i+1}. {name}")
    wb.close()


def preview_sheet(xlsx_path, sheet_name=None):
    """Preview headers and first 5 rows of a sheet."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]
        sheet_name = wb.sheetnames[0]

    print(f"\nPreview of sheet '{sheet_name}':")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    print()

    row_count = 0
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row)):
        row_count += 1
        cells = []
        for cell in row:
            if not hasattr(cell, 'column') or cell.column is None:
                continue
            if cell.value:
                cells.append(f"Col{cell.column}={str(cell.value)[:60]}")
        if cells:
            print(f"  Row {row_count}: {' | '.join(cells)}")
        else:
            print(f"  Row {row_count}: (empty)")

    # Try auto-detection
    print(f"\n  Auto-detection result:")
    header_row, raw_headers, role_map = find_header_row(ws)
    if header_row:
        print(f"    Header row: {header_row}")
        for role, col_num in sorted(role_map.items(), key=lambda x: x[1]):
            print(f"    {role:20s} -> Col {col_num} ({raw_headers.get(col_num, '?')})")
    else:
        print(f"    Could not auto-detect headers.")

    wb.close()


def main():
    parser = argparse.ArgumentParser(
        description='Extract OWASP security rules from Excel file (auto-detect format)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python extract-owasp-rules.py security_rules.xlsx
  python extract-owasp-rules.py rules.xlsx --sheet "Sheet1" --output ./checklist.md
  python extract-owasp-rules.py rules.xlsx --list-sheets
  python extract-owasp-rules.py rules.xlsx --preview
  python extract-owasp-rules.py rules.xlsx --pic All BO Server
        """
    )
    parser.add_argument('xlsx_path', help='Path to the Excel file (.xlsx)')
    parser.add_argument('--output', '-o', help='Output file path (default: owasp_checklist.md in same dir)')
    parser.add_argument('--sheet', '-s', help='Sheet name (default: first sheet)')
    parser.add_argument('--pic', nargs='+', default=['All', 'BO'],
                        help='PIC values to filter (default: All BO)')
    parser.add_argument('--list-sheets', action='store_true',
                        help='List all sheet names and exit')
    parser.add_argument('--preview', action='store_true',
                        help='Preview sheet structure and auto-detection, then exit')

    args = parser.parse_args()

    if not os.path.exists(args.xlsx_path):
        print(f"ERROR: File not found: {args.xlsx_path}", file=sys.stderr)
        sys.exit(1)

    # List sheets mode
    if args.list_sheets:
        list_sheets(args.xlsx_path)
        return

    # Preview mode
    if args.preview:
        preview_sheet(args.xlsx_path, args.sheet)
        return

    # Extract rules
    print(f"Reading: {os.path.basename(args.xlsx_path)}")
    print(f"PIC filter: {args.pic}")
    rules, sheet_name, role_map, total_rows = extract_rules(args.xlsx_path, args.sheet, args.pic)
    print(f"\n  Total rows in sheet: {total_rows}")
    print(f"  After PIC filter: {len(rules)} rules")

    if len(rules) == 0:
        print(f"\nWARNING: No rules matched the filter. Check --pic values or use --preview.", file=sys.stderr)
        sys.exit(1)

    # Generate markdown
    markdown = generate_markdown(rules, sheet_name, args.xlsx_path, role_map)

    # Write output
    if args.output:
        output_path = args.output
    else:
        output_dir = os.path.dirname(os.path.abspath(args.xlsx_path))
        output_path = os.path.join(output_dir, 'owasp_checklist.md')

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(markdown)

    print(f"\nOutput: {output_path}")
    print(f"Done! ({len(rules)} rules extracted)")


if __name__ == '__main__':
    main()
